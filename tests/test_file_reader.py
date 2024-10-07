import unittest
from unittest.mock import mock_open, patch
import csv
import openpyxl
from typing import List, Dict

from file_reader import load_csv, load_xlsx


class TestTransactionLoaders(unittest.TestCase):

    def test_load_csv(self):
            mock_csv_data = """id;state;date;amount;currency_name;currency_code;from;to;description
        1;CANCELED;2021-01-01;100;USD;840;Alice;Bob;Payment
        2;CANCELED;2021-01-02;200;USD;840;Alice;;Transfer
        4;CANCELED;2021-01-04;300;USD;840;;Bob;Invoice"""

            with patch("builtins.open", mock_open(read_data=mock_csv_data)):
                transactions = load_csv('mock_file.csv')
                expected_transactions = [
                    {
                        "id": 1,
                        "state": "CANCELED",
                        "date": "2021-01-01",
                        "operationAmount": {
                            "amount": "100",
                            "currency": {
                                "name": "USD",
                                "code": "840"
                            }
                        },
                        "from": "Alice",
                        "to": "Bob",
                        "description": "Payment"
                    },
                    {
                        "id": 2,
                        "state": "CANCELED",
                        "date": "2021-01-02",
                        "operationAmount": {
                            "amount": "200",
                            "currency": {
                                "name": "USD",
                                "code": "840"
                            }
                        },
                        "from": "Alice",
                        "to": "",
                        "description": "Transfer"
                    },
                    {
                        "id": 4,
                        "state": "CANCELED",
                        "date": "2021-01-04",
                        "operationAmount": {
                            "amount": "300",
                            "currency": {
                                "name": "USD",
                                "code": "840"
                            }
                        },
                        "from": "",
                        "to": "Bob",
                        "description": "Invoice"
                    }
                ]

                self.assertEqual(transactions, expected_transactions)

    def test_load_xlsx(self):
        mock_xlsx_data = [
            ("id", "state", "date", "amount", "currency_name", "currency_code", "from", "to", "description"),
            (1, "CANCELED", "2021-01-01", 100, "USD", 840, "Alice", "Bob", "Payment"),
            (2, "CANCELED", "2021-01-02", 200, "USD", 840, "Alice", "", "Transfer"),
            (4, "CANCELED", "2021-01-04", 300, "USD", 840, "", "Bob", "Invoice"),
        ]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in mock_xlsx_data:
            sheet.append(row)

        with patch("openpyxl.load_workbook") as mock_load_workbook:
            mock_load_workbook.return_value = workbook
            transactions = load_xlsx('mock_file.xlsx')
            expected_transactions = [
                {
                    "id": 1,
                    "state": "CANCELED",
                    "date": "2021-01-01",
                    "operationAmount": {
                        "amount": 100,
                        "currency": {
                            "name": "USD",
                            "code": 840
                        }
                    },
                    "from": "Alice",
                    "to": "Bob",
                    "description": "Payment"
                },
                {
                    "id": 2,
                    "state": "CANCELED",
                    "date": "2021-01-02",
                    "operationAmount": {
                        "amount": 200,
                        "currency": {
                            "name": "USD",
                            "code": 840
                        }
                    },
                    "from": "Alice",
                    "to": "",  # Изменено с None на пустую строку
                    "description": "Transfer"
                },
                {
                    "id": 4,
                    "state": "CANCELED",
                    "date": "2021-01-04",
                    "operationAmount": {
                        "amount": 300,
                        "currency": {
                            "name": "USD",
                            "code": 840
                        }
                    },
                    "from": "",  # Изменено с None на пустую строку
                    "to": "Bob",
                    "description": "Invoice"
                }
            ]
            self.assertEqual(transactions, expected_transactions)

    def test_load_csv_file_not_found(self):
        with patch("builtins.open", side_effect=FileNotFoundError):
            transactions = load_csv('non_existent_file.csv')
            self.assertEqual(transactions, [])

    def test_load_xlsx_file_not_found(self):
        with patch("openpyxl.load_workbook", side_effect=FileNotFoundError):
            transactions = load_xlsx('non_existent_file.xlsx')
            self.assertEqual(transactions, [])
